"""Build docs/rules-review-v2.xlsx — every codec rule with all the
decision-support columns Scott asked for.

Columns:
  # | Section | Stick role | Profile | When | What we emit | 100%? |
  Specific question | Affected (corpus) | Unlocks | My uncertainty (long) |
  Scott's correction (template) | Diff PDF | Manufacturing PDF |
  Architectural PDF | .fcp file | Source XML

Hyperlinks use HYPERLINK() formula so they survive across Excel reopens.
All bare "#N" rule references in text are written as "rule N" so they don't
get auto-converted to GitHub issue links by GitHub's web UI.
"""
import json, os, re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
PDFS = os.path.join(ROOT, "docs", "rule-pdfs")

with open(os.path.join(PDFS, "_links.json"), encoding="utf-8") as f:
    LINKS = json.load(f)
with open(os.path.join(PDFS, "_arch_pdfs.json"), encoding="utf-8") as f:
    ARCH = json.load(f)
with open(os.path.join(PDFS, "_affected.json"), encoding="utf-8") as f:
    AFFECTED = json.load(f)
empirical_path = os.path.join(PDFS, "_empirical_findings.json")
EMPIRICAL = {}
if os.path.exists(empirical_path):
    with open(empirical_path, encoding="utf-8") as f:
        raw = json.load(f)
    EMPIRICAL = {int(k): v for k, v in raw.items()}


def empirical_text(rule_num):
    """Format empirical finding for the spreadsheet cell."""
    fnd = EMPIRICAL.get(rule_num)
    if not fnd:
        return ""
    lines = [f"MINED FROM {fnd['total_sticks']:,} STICKS — overall {fnd['overall_pct']}% match",
             f'"{fnd["headline"]}"',
             ""]
    # Show first 1-2 cuts, top 5 rows each
    for cut in fnd["cuts"][:2]:
        if not cut["rows"]:
            continue
        lines.append(f"By {cut['cut_by']}:")
        for row in cut["rows"][:5]:
            lines.append(f"  {row['key']:<32} {row['with']:>5}/{row['total']:<5} = {row['pct']:>5.1f}%")
        lines.append("")
    return "\n".join(lines).strip()

# (rule_num) -> (frame_name, stick_name) — must match the EXAMPLES in
# render_rule_pdfs.py so the diff PDF and the lookup label refer to the
# same frame/stick.
# Rules whose answer was found in HYTEK config / docs (move from "No" -> "Spec'd").
# Each entry: (source_file_or_concept, what to implement in the codec).
SPEC_FOUND = {
    5: ("docs/service-z-line-design.md",
        "Read <tool_action name=Service> elements per frame. For each horizontal z-line (same z, varying x or y), project onto each stud. Emit InnerService at local_pos = z_line - z_stud_start - 2mm. See doc §1.2 + §3 for selection rule."),
    6: ("docs/service-z-line-design.md",
        "Same per-stud z-line projection as rule 5 — paired hole follows automatically when 2 horizontal z-lines (typically z=300 + z=450) cover the same stud."),
    9: ("machine-setups.ts: chamferTolerance + automaticChamfer",
        "Replace 'Chamfer @start ONLY' with auto-chamfer collision check per-end. Fires when stick end protrudes into adjacent stick within chamferTolerance (4mm in HYTEK setup). Use chamferDetail triangle (14.5×11.5mm) for the cut. endToEndChamfers=false so only protruding end gets chamfered."),
    12: ("machine-setups.ts: chamferDetail + chamferTolerance",
        "Replace '45/cos(angle)' formula with cap span derived from the chamfer triangle's max-x (14.5mm) projected onto the angled stick axis. Once the chamfer cuts, the end-cap follows the cut edge."),
    14: ("machine-setups.ts: chamferTolerance + chamferDetail (89mm setup)",
        "Same auto-chamfer rule as rule 9, applied to 89mm Kb. Setup 6 (F325iT 89mm) has identical chamfer constants to setup 2 (70mm)."),
    29: ("docs/service-z-line-design.md",
        "Same z-line projection rule as rule 5, applied to top plates. Existing post-decode block at scripts/diff-vs-detailer.mjs:642 already does this for T+N nogs — extends to all wall-plan T plates."),
    50: ("machine-setups.ts: setup 6 (89mm) constants identical to setup 2 (70mm) for the relevant fields",
        "Confirmed mirror of 70mm rule. 162-168mm length bucket applies to 89mm too. Already implemented; just verify against 89mm corpus."),
    61: ("machine-setups.ts: chamferTolerance + automaticChamfer",
        "Same auto-chamfer geometric check as rule 9. Replace 'angle >= 28 deg' threshold with collision detection. Wall braces with truly steep angles will fire chamfer; near-vertical wall studs won't."),
    64: ("machine-setups.ts: chamferDetail",
        "Replace '39/cos(a) + 8*tan^2(a)' empirical curve with cap-span derived from the chamfered cut edge. End Swage span = chamferDetail.maxX projected onto stick axis + endClearance."),
    66: ("machine-setups.ts: chamferTolerance + automaticChamfer",
        "Same as rule 61 — auto-chamfer per-end."),
    86: ("fc-dat-rules.ts: cenCenWbHole + bcDistCenHole + webIncrements + horizChdCenHole (per-profile)",
        "Implement panel-point detector: for each Web stick endpoint, project onto the chord centerline. Emit paired InnerDimples at +/- bcDistCenHole (50mm for 70mm normal-axis, 45mm for 89mm) and a LipNotch perpendicular to the chord. Suppress if two panel-points are within cenCenWbHole=120mm."),
    89: ("docs/tb2b-patch-audit.md (14 patches catalogued)",
        "Port 14 patches from scripts/diff-vs-detailer.mjs to src/. Each is documented in tb2b-patch-audit.md with line refs (Web@pt rewrite, box-piece InnerDimple, R-rail short-cap, H4/H7 cap-stacks, T-chord end-cap bolts). Currently TB2B is at 1.34% production parity, 82% with these patches in the harness."),
    90: ("fc-dat-rules.ts: profile-2 (Linear) constants",
        "Linear-truss profile has shortenDblesWb=-50, endWbSetback=50, kpTruncated=90 (vs 0 for STD). Rewire src/simplify-linear-truss.ts to read these from fc-dat-rules.ts instead of hardcoded values."),
    91: ("docs/service-z-line-design.md",
        "Replace blind 'fire InnerService on every wall stud >=500mm' with per-stud z-line projection. Bimodal behaviour disappears when each stud only emits ops for the lines that actually cross it."),
}

EXAMPLE_LOC = {
    5: ("L4", "S11"),
    6: ("L4", "S11"),
    7: ("L4", "S11"),
    8: ("L1101", "S1"),
    12: ("PK4-L4", "Kb1"),
    14: ("L24", "Kb2"),
    15: ("PK4-L4", "H1"),
    19: ("PK4-L4", "T2"),
    20: ("PK4-L4", "T2"),
    28: ("PK4-L4", "T1"),
    29: ("PK4-L4", "T1"),
    30: ("L1101", "T1"),
    32: ("PK1-N1", "B1"),
    33: ("PK1-N1", "B1"),
    34: ("PK1-N1", "B1"),
    36: ("PK1-N1", "B1"),
    38: ("N3", "B2"),
    42: ("PK1-N14", "B1"),
    43: ("PK1-N14", "B1"),
    50: ("L1001", "N1"),
    54: ("PK4-L4", "H1"),
    55: ("PK4-L4", "H1"),
    59: ("PK4-L4", "H1"),
    60: ("L1001", "H1"),
    61: ("PK4-L27", "W2"),
    64: ("PK4-L27", "W2"),
    66: ("PK4-L27", "W2"),
    74: ("L24", "L3"),
    75: ("L24", "L3"),
    79: ("PC2-1", "R3"),
    80: ("PC2-1", "R3"),
    81: ("PC2-1", "R3"),
    82: ("PC2-1", "R3"),
    83: ("R1", "T1"),
    84: ("PK4-L4", "S5"),
    85: ("PK4-L4", "N1"),
    86: ("PC1-1", "T2"),
    87: ("R1", "T1"),
    89: ("TN6-6", "T4"),
    90: ("R1", "T1"),
    91: ("L39", "S7"),
}

# ----- per-rule data -----
# (#, section, role, profile, when, what_we_emit, sure?, example_job, example_plan,
#  arch_pdf_key, specific_question, unlocks)
ROWS = [
    (1, "A", "Wall studs (S, J)", "70mm", "Always", "Start Swage 0..39mm (39mm cap from start)", "Yes", None, None, None, "", "—"),
    (2, "A", "Wall studs (S, J)", "70mm", "Always", "Start InnerDimple at 16.5mm from start", "Yes", None, None, None, "", "—"),
    (3, "A", "Wall studs (S, J)", "70mm", "Always", "End Swage [length-39 .. length]", "Yes", None, None, None, "", "—"),
    (4, "A", "Wall studs (S, J)", "70mm", "Always", "End InnerDimple at length-16.5", "Yes", None, None, None, "", "—"),
    (5, "A", "Wall studs (S, J)", "70/89mm", "Wall plan AND length>=500", "Service hole at 296mm from start (outlet height)", "No", "HG260001", "GF-LBW-70.075", "ac_layout",
        "Which studs get the 296mm hole? Every stud, or only some (e.g. near openings, every Nth, by electrical layout)?",
        "moderate (over-emit on most jobs)"),
    (6, "A", "Wall studs (S, J)", "70/89mm", "Wall plan AND length>=500", "Service hole at 446mm from start (switch height)", "No", "HG260001", "GF-LBW-70.075", "ac_layout",
        "Same discriminator as rule 5. Always paired with the 296 hole?",
        "moderate"),
    (7, "A", "Wall studs (S, J)", "70/89mm", "framecad-import.ts post-process", "Web access holes evenly distributed in body", "No", "HG260001", "GF-LBW-70.075", "carpenter",
        "What's the spacing rule for web access holes (fixed positions vs evenly distributed vs max-gap)?",
        "small"),
    (8, "A", "Wall studs (S, J)", "89mm", "Always", "Same 4-op pattern as 70mm at 89mm constants", "No", "HG260012", "TH01-GF-LBW-89.075", "construction_plans",
        "Do 89mm wall studs really use the same 39mm cap and 16.5mm dimple as 70mm?",
        "small (corpus already aligns)"),

    (9, "B", "Cripple/king studs (Kb)", "70mm", "Kb only", "Chamfer at start ONLY (codec shortcut — actually wrong, real rule is auto-chamfer per-end)", "No", "HG260001", "GF-LBW-70.075", "construction_details",
        "Replaces the always-@start shortcut with auto-chamfer geometric check.",
        "small (covered by auto-chamfer rewrite)"),
    (10, "B", "Cripple/king studs (Kb)", "70mm", "Kb only", "Start Swage 0..42mm", "Yes", None, None, None, "", "—"),
    (11, "B", "Cripple/king studs (Kb)", "70mm", "Kb only", "Start InnerDimple at 10mm (NOT 16.5)", "Yes", None, None, None, "", "—"),
    (12, "B", "Cripple/king studs (Kb)", "70mm", "Kb only, angle-dependent", "End Swage span = 45 / cos(angle from horizontal)", "No", "HG260001", "GF-LBW-70.075", "construction_details",
        "What's the actual rule for the Kb end-Swage span — is 45mm a HYTEK constant?",
        "small"),
    (13, "B", "Cripple/king studs (Kb)", "70mm", "Kb only", "End InnerDimple at length-10", "Yes", None, None, None, "", "—"),
    (14, "B", "Cripple/king studs (Kb)", "89mm", "Always", "Same Kb pattern as 70mm", "No", "HG250030", "GF-LBW-89.075", None,
        "Do 89mm Kb sticks use the same 42mm + 45/cos pattern as 70mm?",
        "small"),

    (15, "C", "Header stud (H, single)", "70mm", "H not Kb", "End Swage span 43mm", "No", "HG260001", "GF-LBW-70.075", "construction_details",
        "Is the cripple-role H rule actually live, or dead code? When does H differ from Kb?",
        "negligible"),

    (16, "D", "Short top plate (T<200mm)", "70mm", "Length<200mm", "Start InnerNotch span 39mm", "Yes", None, None, None, "", "—"),
    (17, "D", "Short top plate (T<200mm)", "70mm", "Length<200mm", "Start LipNotch span 39mm", "Yes", None, None, None, "", "—"),
    (18, "D", "Short top plate (T<200mm)", "70mm", "Length<200mm", "Start InnerDimple at 16.5", "Yes", None, None, None, "", "—"),
    (19, "D", "Short top plate (T<200mm)", "70mm", "Length<200mm AND LBW", "Start InnerDimple at 58.5 (paired)", "No", "HG260001", "GF-LBW-70.075", "lbw_markup",
        "Why is the 58.5 paired dimple LBW-only? What does it represent physically (strap, screw, joist hanger)?",
        "small"),
    (20, "D", "Short top plate (T<200mm)", "70mm", "Length<200mm AND LBW", "End InnerDimple at length-58.5", "No", "HG260001", "GF-LBW-70.075", "lbw_markup",
        "Same answer as rule 19 (mirror at end).",
        "small"),
    (21, "D", "Short top plate (T<200mm)", "70mm", "Length<200mm", "End InnerDimple at length-16.5", "Yes", None, None, None, "", "—"),
    (22, "D", "Short top plate (T<200mm)", "70mm", "Length<200mm", "End InnerNotch span 39mm", "Yes", None, None, None, "", "—"),
    (23, "D", "Short top plate (T<200mm)", "70mm", "Length<200mm", "End LipNotch span 39mm", "Yes", None, None, None, "", "—"),

    (24, "E", "Long top plate (T>=200mm)", "70mm", "Length>=200mm", "Start LipNotch span 39mm", "Yes", None, None, None, "", "—"),
    (25, "E", "Long top plate (T>=200mm)", "70mm", "Length>=200mm", "Start InnerDimple at 16.5", "Yes", None, None, None, "", "—"),
    (26, "E", "Long top plate (T>=200mm)", "70mm", "Length>=200mm", "End LipNotch span 39mm", "Yes", None, None, None, "", "—"),
    (27, "E", "Long top plate (T>=200mm)", "70mm", "Length>=200mm", "End InnerDimple at length-16.5", "Yes", None, None, None, "", "—"),
    (28, "E", "Long top plate (T>=200mm)", "70mm", "DELIBERATELY NOT EMITTED", "InnerNotch on long T plates", "No", "HG260001", "GF-LBW-70.075", "construction_details",
        "When SHOULD a long T plate get InnerNotch? At king-stud crossings only? Above openings? At plate joints?",
        "moderate"),
    (29, "E", "Long top plate (T>=200mm)", "70mm", "DELIBERATELY NOT EMITTED", "InnerService holes on T plates", "No", "HG260001", "GF-LBW-70.075", "ac_layout",
        "Which T plates get service holes? Same electrical-data discriminator as rule 5/6?",
        "moderate"),
    (30, "E", "Long top plate (T>=200mm)", "89mm", "Length>=200mm", "Same 4-op pattern as 70mm", "No", "HG260012", "TH01-GF-LBW-89.075", "construction_plans",
        "Do 89mm long T plates match 70mm pattern?",
        "small"),

    (31, "F", "Bottom plate (B, Bp)", "70mm", "Always", "Start LipNotch span 39mm", "Yes", None, None, None, "", "—"),
    (32, "F", "Bottom plate (B, Bp)", "70mm", "Primary B + GF + LBW/NLBW", "Start Web hole at 8mm (slab wiring)", "No", "HG260001", "GF-LBW-70.075", "slab_design",
        "Confirm gates: primary B (B1 or >=1500mm) + ground floor + wall plan?",
        "moderate"),
    (33, "F", "Bottom plate (B, Bp)", "70mm", "Always", "Start InnerDimple at 16.5", "Yes", None, None, None, "", "—"),
    (34, "F", "Bottom plate (B, Bp)", "70mm", "Same gates as rule 32", "Start Bolt at 62mm (slab anchor)", "No", "HG260001", "GF-LBW-70.075", "slab_design",
        "Where does the 62mm offset come from? Always 62 or scales with profile/bolt spec? Middle bolts on long plates?",
        "moderate"),
    (35, "F", "Bottom plate (B, Bp)", "70mm", "Always", "End LipNotch span 39mm", "Yes", None, None, None, "", "—"),
    (36, "F", "Bottom plate (B, Bp)", "70mm", "Same gates as rule 32", "End Bolt at length-62", "No", "HG260001", "GF-LBW-70.075", "slab_design",
        "Same as rule 34 (mirror at end).",
        "moderate"),
    (37, "F", "Bottom plate (B, Bp)", "70mm", "Always", "End InnerDimple at length-16.5", "Yes", None, None, None, "", "—"),
    (38, "F", "Bottom plate (B, Bp)", "89mm", "Same gates as rule 32 + gauge<1.0", "Slab anchor pattern (Web@8 + Bolts@62)", "No", "HG250082", "GF-NLBW-89.075", "slab_design",
        "Do thicker 89mm B plates (gauge >=1.0) really skip slab bolts in real production?",
        "small"),

    (39, "G", "Raised B plate (Bh)", "70mm", "Always", "Start InnerNotch span 39mm", "Yes", None, None, None, "", "—"),
    (40, "G", "Raised B plate (Bh)", "70mm", "Always", "Start LipNotch span 39mm", "Yes", None, None, None, "", "—"),
    (41, "G", "Raised B plate (Bh)", "70mm", "Always", "Start InnerDimple at 16.5", "Yes", None, None, None, "", "—"),
    (42, "G", "Raised B plate (Bh)", "70mm", "NLBW only", "Start Web@8 + Bolt@62 (slab anchor on raised B)", "No", "HG260001", "GF-NLBW-70.075", "slab_design",
        "WHY does NLBW raised B get slab anchors but LBW raised B doesn't?",
        "small"),
    (43, "G", "Raised B plate (Bh)", "70mm", "NLBW only", "End Bolt at length-62", "No", "HG260001", "GF-NLBW-70.075", "slab_design",
        "Same as rule 42 (mirror at end).",
        "small"),
    (44, "G", "Raised B plate (Bh)", "70mm", "Always", "End InnerDimple at length-16.5", "Yes", None, None, None, "", "—"),
    (45, "G", "Raised B plate (Bh)", "70mm", "Always", "End InnerNotch span 39mm", "Yes", None, None, None, "", "—"),
    (46, "G", "Raised B plate (Bh)", "70mm", "Always", "End LipNotch span 39mm", "Yes", None, None, None, "", "—"),
    (47, "G", "Raised B plate (Bh)", "89mm", "Always", "Header-style caps both ends, NO Web/Bolt", "Yes", None, None, None, "", "—"),

    (48, "H", "Nogs (N) — door-head cripple", "70mm", "Length 162-168mm", "InnerNotch + LipNotch caps + Dimple@16.5 BOTH ends", "Yes", None, None, None, "", "—"),
    (49, "H", "Nogs (N) — all other lengths", "70mm", "Length NOT in 162-168mm", "Swage caps + Dimple@16.5 BOTH ends", "Yes", None, None, None, "", "—"),
    (50, "H", "Nogs (N)", "89mm", "Always (length-bucketed)", "Same pattern as 70mm with 89mm constants", "No", "HG260012", "TH01-GF-LBW-89.075", "construction_plans",
        "Confirm 89mm nog uses same 162-168mm length bucket and same caps as 70mm.",
        "small"),

    (51, "I", "Header (H — full)", "70mm", "Always", "Start InnerNotch span 39mm", "Yes", None, None, None, "", "—"),
    (52, "I", "Header (H — full)", "70mm", "Always", "Start LipNotch span 39mm", "Yes", None, None, None, "", "—"),
    (53, "I", "Header (H — full)", "70mm", "Always", "Start InnerDimple at 16.5", "Yes", None, None, None, "", "—"),
    (54, "I", "Header (H — full)", "70mm", "LBW only", "Start InnerDimple at 58.5 (paired)", "No", "HG260001", "GF-LBW-70.075", "lbw_markup",
        "Same LBW-only question as rule 19. Why LBW-specific?",
        "small"),
    (55, "I", "Header (H — full)", "70mm", "LBW only", "End InnerDimple at length-58.5", "No", "HG260001", "GF-LBW-70.075", "lbw_markup",
        "Same as rule 54 (mirror at end).",
        "small"),
    (56, "I", "Header (H — full)", "70mm", "Always", "End InnerDimple at length-16.5", "Yes", None, None, None, "", "—"),
    (57, "I", "Header (H — full)", "70mm", "Always", "End InnerNotch span 39mm", "Yes", None, None, None, "", "—"),
    (58, "I", "Header (H — full)", "70mm", "Always", "End LipNotch span 39mm", "Yes", None, None, None, "", "—"),
    (59, "I", "Header (H — full)", "70mm", "H1 or H3 in paired-header frame", "Web stiffener holes evenly distributed", "No", "HG260001", "GF-LBW-70.075", "construction_details",
        "What physically makes a header 'paired' (box vs single-section)? When does the rule fire?",
        "moderate"),
    (60, "I", "Header (H — full)", "89mm", "Always", "Caps + paired dimples at 16.5 AND 58.5 always both ends", "No", "HG260012", "TH01-GF-LBW-89.075", "construction_plans",
        "Do 89mm NLBW headers also get the paired 58.5 dimple?",
        "small"),

    (61, "J", "Wall brace (W on wall plan)", "70/89mm", "Wall plan AND angle from vertical >= 28 deg", "Chamfer at start (diagonal cut)", "No", "HG260001", "GF-LBW-70.075", "construction_details",
        "Is 28 deg the actual chamfer threshold? What's the real rule?",
        "small"),
    (62, "J", "Wall brace (W on wall plan)", "70/89mm", "Wall plan", "Start Swage span 41mm", "Yes", None, None, None, "", "—"),
    (63, "J", "Wall brace (W on wall plan)", "70/89mm", "Wall plan", "Start InnerDimple at 10mm", "Yes", None, None, None, "", "—"),
    (64, "J", "Wall brace (W on wall plan)", "70/89mm", "Wall plan, angle-dependent", "End Swage span = 39/cos(a) + 8*tan^2(a)", "No", "HG260001", "GF-LBW-70.075", "construction_details",
        "What's the actual rule for end-Swage span on a wall brace? Fixed perpendicular cap, or angle-scaled formula?",
        "moderate"),
    (65, "J", "Wall brace (W on wall plan)", "70/89mm", "Wall plan", "End InnerDimple at length-10", "Yes", None, None, None, "", "—"),
    (66, "J", "Wall brace (W on wall plan)", "70/89mm", "Wall plan AND angle >= 28 deg", "End Chamfer", "No", "HG260001", "GF-LBW-70.075", "construction_details",
        "Same threshold as rule 61.",
        "small"),

    (67, "K", "Truss/joist web (W non-wall)", "70/89mm", "Non-wall plan", "Start Swage span 39mm", "Yes", None, None, None, "", "—"),
    (68, "K", "Truss/joist web (W non-wall)", "70/89mm", "Non-wall plan", "Start InnerDimple at 16.5", "Yes", None, None, None, "", "—"),
    (69, "K", "Truss/joist web (W non-wall)", "70/89mm", "Non-wall plan", "End Swage span 39mm", "Yes", None, None, None, "", "—"),
    (70, "K", "Truss/joist web (W non-wall)", "70/89mm", "Non-wall plan", "End InnerDimple at length-16.5", "Yes", None, None, None, "", "—"),

    (71, "L", "Lintel/sill (L)", "70mm", "Always", "Start InnerNotch span 39mm", "Yes", None, None, None, "", "—"),
    (72, "L", "Lintel/sill (L)", "70mm", "Always", "Start LipNotch span 39mm", "Yes", None, None, None, "", "—"),
    (73, "L", "Lintel/sill (L)", "70mm", "Always", "Start InnerDimple at 16.5", "Yes", None, None, None, "", "—"),
    (74, "L", "Lintel/sill (L)", "89mm", "Always", "Start InnerDimple at 58.5 (paired)", "No", "HG260002", "GF-LBW-89.075", "construction_details",
        "Why does 89mm L always get the 58.5 dimple but 70mm L doesn't? Or is it actually LBW-only too?",
        "small"),
    (75, "L", "Lintel/sill (L)", "89mm", "Always", "End InnerDimple at length-58.5", "No", "HG260002", "GF-LBW-89.075", "construction_details",
        "Same as rule 74.",
        "small"),
    (76, "L", "Lintel/sill (L)", "70mm", "Always", "End InnerDimple at length-16.5", "Yes", None, None, None, "", "—"),
    (77, "L", "Lintel/sill (L)", "70mm", "Always", "End InnerNotch span 39mm", "Yes", None, None, None, "", "—"),
    (78, "L", "Lintel/sill (L)", "70mm", "Always", "End LipNotch span 39mm", "Yes", None, None, None, "", "—"),

    (79, "M", "Brace/Ribbon (Br, R)", "70mm", "Always", "Start Swage span 41mm", "No", "HG260044", "GF-TIN-70.075", "construction_details",
        "What's the difference between Br, R, and W? Should Br/R use 41mm cap + 11mm dimple or same as studs?",
        "small"),
    (80, "M", "Brace/Ribbon (Br, R)", "70mm", "Always", "Start InnerDimple at 11mm", "No", "HG260044", "GF-TIN-70.075", "construction_details",
        "Same as rule 79.", "small"),
    (81, "M", "Brace/Ribbon (Br, R)", "70mm", "Always", "End Swage span 41mm", "No", "HG260044", "GF-TIN-70.075", "construction_details",
        "Same as rule 79.", "small"),
    (82, "M", "Brace/Ribbon (Br, R)", "70mm", "Always", "End InnerDimple at length-11", "No", "HG260044", "GF-TIN-70.075", "construction_details",
        "Same as rule 80.", "small"),

    (83, "N", "Crossings: plate over stud", "All", "Stud crosses plate", "Plate gets LipNotch + InnerDimple at every stud's x-coord", "No", "HG260002", "GF-RP-70.075", "construction_details",
        "At a stud-plate crossing, what tool does Detailer emit on the plate (LipNotch / InnerNotch / Web)? What's the discriminator?",
        "moderate"),
    (84, "N", "Crossings: stud over nog", "All", "Nog crosses stud", "Stud gets LipNotch + InnerDimple at every nog's y-coord", "No", "HG260001", "GF-LBW-70.075", "construction_details",
        "At a nog-stud crossing, what tool does Detailer emit on the stud?",
        "small"),
    (85, "N", "Crossings: nog over stud", "All", "Stud crosses nog", "Nog gets Web + LipNotch + InnerDimple at every stud's x-coord", "No", "HG260001", "GF-LBW-70.075", "construction_details",
        "At a stud-nog crossing, do all three (Web + LipNotch + InnerDimple) always fire?",
        "small"),
    (86, "N", "Crossings: truss panel-points", "All", "Web stick projects onto chord", "NOT EMITTED — paired InnerDimples + 47mm-wide LipNotch", "No", "HG250011", "GF-TIN-70.075", "construction_details",
        "Are 51mm spacing and 47mm LipNotch width HYTEK constants? Offset from Web projection always 41mm before?",
        "+13pp on TIN (BIGGEST single gap)"),

    (87, "O", "Simplifier: simplify-rp.ts", "All", "-RP- plans", "Rewrites caps to chord-style on T/B/N + plate-over-plate on S", "No", "HG260006", "GF-RP-70.075", "construction_details",
        "What discriminates an RP frame that gets chord-style caps vs one that gets stud-style caps?",
        "moderate (~+8pp on RP)"),
    (88, "O", "Simplifier: simplify-tin-truss.ts", "All", "-TIN- plans", "H-stick LipNotch->Swage substitution", "Yes", None, None, None, "", "—"),
    (89, "O", "Simplifier: simplify-tb2b-truss.ts", "All", "-TB2B- + frame=Truss", "TB2B truss frame cap rewrites", "No", "HG260001", "GF-TB2B-70.075", "construction_details",
        "What's the TB2B truss tooling pattern at scale? (Need broader corpus check first.)",
        "unknown"),
    (90, "O", "Simplifier: simplify-linear-truss.ts", "All", "-LIN- plans", "Bolt-hole pattern at every web-chord crossing", "No", "HG260043", "GF-RP-89.115", "construction_details",
        "What are LIN plans for? Bolt-hole spacing rule?",
        "unknown"),
    (91, "O", "Simplifier: simplify-wall-service.ts", "All", "-LBW-/-NLBW- plans", "InnerService holes on wall studs", "No", "HG260010", "GF-LBW-70.075", "ac_layout",
        "Same as rule 5/6 — what tells Detailer WHICH studs need service holes?",
        "moderate"),
]

# ----- long-form uncertainty (replaces the v1 LONG dict; uses 'rule N' not '#N') -----
LONG = {
    5: "WHAT WE EMIT: An InnerService hole at exactly 296mm from the BOTTOM of every wall stud (S/J), but only when the plan is LBW or NLBW and the stud is at least 500mm long. Position is fixed.\n\n"
       "WHAT DETAILER DOES: Detailer emits InnerService holes on wall studs at this height too — but it is SELECTIVE. It only puts the hole on SOME of the studs, not all. When we emit on every stud, we generate dozens of extra holes Detailer didn't put there.\n\n"
       "MY HYPOTHESIS: Detailer is reading something from project metadata that we don't import — most likely electrical layout data (which studs are next to outlets, switches, light fittings). Or it's picking studs based on a fixed counting pattern (every 3rd stud, or only studs adjacent to openings).\n\n"
       "WHAT I NEED FROM YOU: How does Detailer (or your standard practice) decide WHICH studs get the 296mm hole?",
    6: "Same shape as rule 5 — paired hole at 446mm. Once we know which studs get the 296 hole, the 446 follows. Also: are the 296 / 446 heights ever different (e.g. taller walls, non-domestic occupancy)?",
    7: "WHAT WE EMIT: A series of evenly-spaced web access holes along the body of every wall stud, generated by framecad-import.ts post-processing. The number and spacing are hand-tuned, not derived from a HYTEK rule.\n\n"
       "WHAT I NEED FROM YOU: For a 2400mm wall stud, what web holes does Detailer emit (positions in mm from the bottom)? Same question for a 2700mm and a 1800mm stud.",
    8: "WHAT WE EMIT: 89mm wall studs use the same 39mm Swage cap + 16.5mm dimple as 70mm. Constants come from MACHINE_SETUPS['6'] (F325iT 89mm) which has Tab=35 + Clearance=4 = 39, identical to 70mm.\n\n"
       "WHAT I NEED FROM YOU: Confirm 89mm wall studs really use the same 39mm cap + 16.5mm dimple. If different, what are the values?",
    12: "WHAT WE EMIT: For Kb sticks the END Swage span scales with the stick angle. Formula: span = 45 / cos(angle from horizontal). Vertical Kb -> 45mm. 68.3 deg from horizontal -> 121.6mm.\n\n"
        "WHAT DETAILER DOES: HG260001 PK4 LBW Kb1 at 68.3 deg has ref Swage 1354..1476.6 (span 122.6mm). Formula predicts 121.6 — match. We have one verified data point.\n\n"
        "WHAT WORRIES ME: 45 isn't a HYTEK constant I can find. The 39mm (tab + clearance) is documented; 45mm just happens to fit.\n\n"
        "WHAT I NEED FROM YOU: Is there a HYTEK / FrameCAD specification for cripple/king stud end caps? Specifically: angle-dependent or fixed? 45mm correct or different? Formal rule?",
    14: "WHAT WE EMIT: 89mm Kb sticks use the same Chamfer + Swage + Dimple pattern as 70mm with the angle formula at the end. Untested at scale on 89mm.\n\n"
        "WHAT I NEED FROM YOU: Same as rule 12 but for 89mm — are constants (45mm, 10mm) the same?",
    15: "WHAT WE EMIT: When a stick named H is matched as 'cripple' (not full header), end Swage span is 43mm.\n\n"
        "WHAT WORRIES ME: There are two rules for sticks named H. Since H sticks always have role='H', HEADER_ROLES always wins, so this 43mm rule is probably DEAD CODE.\n\n"
        "WHAT I NEED FROM YOU: Are there ever sticks named H that should be treated like cripple studs (43mm Swage cap) instead of full headers?",
    19: "WHAT WE EMIT: Short top sub-plates (T<200mm, sit ABOVE a header) get a paired dimple at 58.5mm from the start. ONLY on LBW plans, not NLBW.\n\n"
        "WHAT DETAILER DOES: HG260001 PK4 LBW L4 (T2/T3/T4 sub-plates 121-127mm) shows ref dimples at 58.5. Our NLBW corpus doesn't.\n\n"
        "WHAT I NEED FROM YOU: What does the 58.5 dimple represent physically (strap, header connector, joist hanger)? Why is it LBW-only?",
    20: "Mirror of rule 19 at the end (length-58.5). Once rule 19 is settled, the other follows.",
    28: "WHAT WE EMIT: Nothing — InnerNotch on long top plates is DELIBERATELY DISABLED. Blind emission produced 100 extras vs only 12 matches.\n\n"
        "WHAT DETAILER DOES: SOME long T plates DO get InnerNotches. We just haven't found the discriminator.\n\n"
        "WHAT I NEED FROM YOU: When does a long T plate get an InnerNotch? At king-stud crossings only? Above openings? At plate joints?",
    29: "WHAT WE EMIT: Nothing — InnerService holes on T plates are DELIBERATELY DISABLED. Blind emission produced 256 extras vs 14 missing.\n\n"
        "MY HYPOTHESIS: Same as rule 5/6 — Detailer reads electrical-services data we don't import.\n\n"
        "WHAT I NEED FROM YOU: What tells Detailer to put service holes on a particular top plate?",
    30: "Same shape as rule 8 — 89mm long T plates use the 70mm pattern. Confirm or correct.",
    32: "WHAT WE EMIT: A Web hole at 8mm from the start of a 70mm bottom plate, but only when ALL THREE conditions are true: (a) the plate is 'primary' (B1, OR length >= 1500mm), (b) the plan name has '-GF-' (ground floor), (c) the plan is LBW or NLBW.\n\n"
        "WHAT DETAILER DOES: This is what I observed across 4 corpora — but the rule was assembled by trial-and-error. B-sub-plates above doors, non-ground-floor walls, and truss/joist plans don't get Web@8.\n\n"
        "WHAT I NEED FROM YOU: Confirm the three gates. Is 8mm the right offset? Any 89mm-specific differences?",
    34: "WHAT WE EMIT: A Bolt at 62mm from the start of a primary ground-floor B plate. Same gates as rule 32.\n\n"
        "WHAT WORRIES ME: 62mm isn't documented. The setup field 'boltHoleToEnd' = 20 doesn't match. 62mm is empirical.\n\n"
        "WHAT I NEED FROM YOU: Where does 62mm come from? Always 62 or scales with plate length / bolt spec? For long plates, do you put MORE bolts in the middle?",
    36: "Mirror of rule 34 at the end. Once 34 is settled, this follows. ALSO: when the plate is short, does it still get bolts at both ends or just one centered?",
    38: "WHAT WE EMIT: 89mm bottom plates get the slab-anchor pattern (Web@8 + Bolt@62 each end) ONLY when gauge < 1.0mm.\n\n"
        "WHAT DETAILER DOES: HG260044 GF-NLBW-89.075 (gauge 0.75) has Web@8 + Bolt@~50. Single corpus. We have no 89mm B-plate samples at gauge 1.0+.\n\n"
        "WHAT I NEED FROM YOU: Do thicker (1.0+, 1.15) 89mm plates really skip slab bolts in real production?",
    42: "WHAT WE EMIT: Raised B plates (Bh, the rough opening sill above doors at z=elevation+61.5mm) get slab-anchor ops (Web@8 + Bolt@62) ONLY on NLBW plans. LBW raised B doesn't get them.\n\n"
        "WHAT DETAILER DOES: HG260001 PK1 N14/B1 (NLBW raised B at length 1872, z=61.5) has cap notches AND Web@8 + Bolt@62. PK4 L4/B2 (LBW raised B) only has cap notches.\n\n"
        "WHAT I NEED FROM YOU: WHY does NLBW raised B get slab anchors but LBW raised B doesn't? Is it the load path? Opening type? Wall load class?",
    43: "Mirror of rule 42 at the end.",
    50: "Same shape as rule 8 — 89mm nogs use the 70mm length-bucketed pattern. Confirm 162-168mm bucket is the same; confirm cap pattern matches.",
    54: "Same as rule 19 — full headers (70mm) get the paired 58.5 dimple LBW-only. Why LBW-specific? Same physical reason as the short T sub-plate?",
    55: "Mirror of rule 54 at the end.",
    59: "WHAT WE EMIT: Sticks named H1 or H3 in frames that have a 'paired header' (frame contains H2 or H3 alongside H1) get web stiffener holes evenly distributed. Spacing: 89mm offset from each end, max 300mm between holes.\n\n"
        "WHAT DETAILER DOES: HG260001 LBW H1/H3 corpus (10 frames) match. Single-H frames (just H1, no H2) don't show stiffener holes.\n\n"
        "MY HYPOTHESIS: A 'paired header' is a built-up box header (two C-sections back-to-back) for wider openings or higher loads.\n\n"
        "WHAT I NEED FROM YOU: What physically creates a paired header — span width? Load class? Specific frame type? Architectural detail? Are stiffeners only on H1/H3, or do H2 also get them?",
    60: "WHAT WE EMIT: 89mm headers ALWAYS get the paired 58.5 dimple at both ends, regardless of LBW/NLBW. (70mm only on LBW.)\n\n"
        "WHAT WORRIES ME: My 89mm header sample is small and may all be LBW. The rule might really be 'LBW only' for 89mm too.\n\n"
        "WHAT I NEED FROM YOU: Do 89mm NLBW headers get the paired 58.5 dimple?",
    61: "WHAT WE EMIT: Wall braces get a Chamfer at start ONLY if angle from vertical >= 28 deg. Near-vertical W's (under 28 deg) don't get a chamfer.\n\n"
        "WHAT DETAILER DOES: HG260001 PK4 L27/W6 at 25.48 deg has NO chamfer. L27/W2 at 29.31 deg HAS a chamfer.\n\n"
        "WHAT I NEED FROM YOU: Is 28 deg the actual threshold, or something else (15? 20? 30? 45?)? Is it really an angle threshold or is it triggered by something else (a flag in the XML, the brace type)?",
    64: "WHAT WE EMIT: For wall braces, the END Swage span is angle-dependent. Formula: span = 39 / cos(angle from vertical) + 8 * tan^2(angle). Curve fit, RMSE 0.36mm across 280 W's.\n\n"
        "WHAT WORRIES ME: The formula gives the right NUMBERS but I don't know if it's the right MODEL. The 39/cos part is geometric (perpendicular cap projected onto angled stick); the 8*tan^2 is a residual fit with no physical interpretation.\n\n"
        "WHAT I NEED FROM YOU: What's the actual rule for end Swage span on a wall brace? (a) Detailer literal formula, (b) fixed perpendicular cap and the projection makes it look angle-dependent, (c) cap follows the cut line?",
    66: "Mirror of rule 61 at the end of the brace.",
    74: "WHAT WE EMIT: 89mm sills (L role) get a paired dimple at 58.5 from each end. 70mm lintels DON'T.\n\n"
        "WHAT WORRIES ME: Same shape as the LBW-paired-dimple question. Why does 89mm L always get it but 70mm L doesn't? Is it really profile-dependent?\n\n"
        "WHAT I NEED FROM YOU: Do 70mm LBW lintels get the 58.5 paired dimple too? If so, the rule should match 70mm headers (LBW gets paired, NLBW doesn't). Is 58.5 always exactly that, or does it scale?",
    75: "Mirror of rule 74 at the end.",
    79: "WHAT WE EMIT: Sticks named with prefix Br or R get a Swage cap of 41mm at the start (vs 39mm for studs, 42mm for Kb).\n\n"
        "WHAT WORRIES ME: Single sample subset.\n\n"
        "WHAT I NEED FROM YOU: What's the difference between Br, R, and W in HYTEK terminology? Should Br/R use 41mm cap + 11mm dimple, or actually use 39mm + 16.5mm like studs?",
    80: "Same as rule 79 — confirm 11mm dimple offset or correct.",
    81: "Same as rule 79 — mirror at end.",
    82: "Same as rule 80 — mirror at end.",
    83: "WHAT WE EMIT: Where a stud crosses a top or bottom plate, the plate gets a LipNotch (39mm spanned) + InnerDimple at the stud's x-coordinate. Position is computed from each stick's outline corners (the elevation graphics in the XML).\n\n"
        "WHAT DETAILER DOES: Position is mostly RIGHT (within ~5mm). The TOOL TYPE is sometimes wrong on RP plans — Detailer uses InnerNotch where we use LipNotch, or Web where we use neither.\n\n"
        "WHAT I NEED FROM YOU: At a stud-plate crossing, what tool does Detailer emit and what determines the type? Always LipNotch on T/B? Or does the stud type (Kb vs S) flip it? Plan-type-specific?",
    84: "Same kind of issue as rule 83 — at a nog-stud crossing, what tool does Detailer put on the stud?",
    85: "Same kind of issue as rule 83 — at a stud-nog crossing, do all three (Web + LipNotch + InnerDimple) always fire, or do some plans/configurations skip?",
    86: "WHAT WE EMIT: NOTHING. We currently emit no ops at truss panel-points.\n\n"
        "WHAT DETAILER DOES: At every panel point (where a Web stick projects onto the chord centerline), Detailer emits paired InnerDimples spaced ~51mm apart, plus a 47mm-wide LipNotch perpendicular to the chord.\n\n"
        "SCALE OF THE GAP: ~7,500 missing ops across the 90 TIN pairs. Worth approximately +13pp on TIN-only parity.\n\n"
        "WHAT I NEED FROM YOU: Is the 51mm spacing a HYTEK constant or panel-spacing dependent? Is the LipNotch always 47mm wide? Offset from Web projection always 41mm before? Does Detailer also emit something on the Web stick at this junction? Does this apply to TB2B and FJ as well as TIN?",
    87: "WHAT WE EMIT: For RP plans, the simplifier rewrites T/B/N plate caps to chord-style (Chamfer + ID@10) and S studs to plate-over-plate notch (Swage 56..101 + ID@78.5).\n\n"
        "WHAT DETAILER DOES: HALF the 50-pair RP corpus uses chord-style caps (matches our simplifier). The OTHER HALF uses STUD-STYLE caps on T/B/N (LipNotch 56..101 + ID@78.5 + ID@170.5...). Same letter, same plan type, different cap regime. Our simplifier helps the first group by 15+pp; HURTS the second by up to 36pp.\n\n"
        "EXAMPLES: HG260001 GF-RP-70.075 has chord-style T plates (we help). HG260006 GF-RP-70.075 has stud-style T plates (we hurt — 33%->69% if disabled).\n\n"
        "MY HYPOTHESES: (a) edge of panel vs middle, (b) single-bay vs multi-bay, (c) frame skew angle, (d) some XML attribute we don't read.\n\n"
        "WHAT I NEED FROM YOU: What physically separates an RP frame that gets chord-style caps from one that gets stud-style?",
    89: "WHAT WE EMIT: TB2B truss frame cap rewrites — hand-tuned on a single job.\n\n"
        "WHAT WORRIES ME: Untested at scale on the 54-pair TB2B corpus.\n\n"
        "WHAT I NEED FROM YOU: Anything you know about TB2B truss tooling that's different from regular truss webs.",
    90: "WHAT WE EMIT: For -LIN- plan-type linear-truss frames, the simplifier emits a bolt-hole pattern at every web-chord centerline crossing.\n\n"
        "WHAT WORRIES ME: No corpus check beyond the original 1-job tuning. Few -LIN- pairs in the corpus.\n\n"
        "WHAT I NEED FROM YOU: What are -LIN- plans for, physically? Bolt-hole spacing rule?",
    91: "WHAT WE EMIT: For LBW/NLBW wall plans, the simplifier emits InnerService holes on every wall stud above a certain length.\n\n"
        "WHAT DETAILER DOES: BIMODAL. On small jobs (HG260010, 170 sticks): Detailer emits 8 service ops total; we emit 323 (~40x over-emit). On big jobs (HG260045): Detailer 73, we emit 55 (under-emit).\n\n"
        "WHAT I NEED FROM YOU: Same as rule 5/6 — what tells Detailer WHICH studs need service holes?",
}

# ----- decision template scaffold (pre-fill in correction column for every No row) -----
TEMPLATE = (
    "THE RULE SHOULD FIRE WHEN: \n"
    "THE VALUES ARE: \n"
    "IT SHOULD NOT FIRE WHEN: \n"
    "WHY (physical reason): "
)

# ----- workbook -----
wb = Workbook()
ws = wb.active
ws.title = "Codec rules"

HEADERS = [
    "#",
    "Section",
    "Stick role",
    "Profile",
    "When (gate / condition)",
    "What we emit (plain English)",
    "Status",
    "Source / what to implement",
    "Empirical pattern (mined from 388-pair corpus)",
    "Look up in Detailer (job / plan / frame / stick)",
    "Specific question",
    "Affected (sticks / pairs in 388-pair corpus)",
    "Unlocks if fixed",
    "My uncertainty (long-form)",
    "Scott's correction",
    "Diff PDF (codec vs Detailer)",
    "Manufacturing PDF (frame elevations)",
    "Architectural PDF (project context)",
    ".fcp file (open in Detailer)",
    "Source XML",
]
ws.append(HEADERS)

font_header = Font(name="Arial", bold=True, color="FFFFFF", size=11)
fill_header = PatternFill("solid", start_color="231F20")
align_header = Alignment(horizontal="center", vertical="center", wrap_text=True)
fill_yes = PatternFill("solid", start_color="C8E6C9")
fill_specd = PatternFill("solid", start_color="FFE0B2")  # orange/amber for Spec'd
fill_no = PatternFill("solid", start_color="FFCDD2")
fill_correction = PatternFill("solid", start_color="FFF9C4")
align_wrap = Alignment(vertical="top", wrap_text=True)
align_center = Alignment(horizontal="center", vertical="center")
font_link = Font(name="Arial", size=10, color="1976D2", underline="single")
thin = Side(border_style="thin", color="CCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for cell in ws[1]:
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = align_header
    cell.border = border
ws.row_dimensions[1].height = 50

# Build job/plan -> manufacturing PDF resolver (LINKS already loaded above)
def manuf_pdf(job, plan):
    p = LINKS.get('manufacturing_pdf', {}).get(job, {}).get(plan)
    if not p:
        p = LINKS.get('site_layout_pdf', {}).get(job)
    return p

def xml_path(job, plan):
    return LINKS.get('xml', {}).get(job, {}).get(plan)

def fcp_path(job):
    return LINKS.get('fcp', {}).get(job)

def arch_pdf_path(job, key):
    if not job or not key:
        return None
    return ARCH.get(job, {}).get(key)

def hyperlink(path, label):
    """Build an Excel HYPERLINK formula. Path is a Windows file:// URI."""
    if not path:
        return ""
    # Excel needs file:// prefix and forward slashes; spaces are OK
    uri = path.replace("\\", "/")
    if not uri.startswith("file:"):
        uri = "file:///" + uri
    # Escape double quotes in path (rare)
    safe_path = uri.replace('"', '""')
    safe_label = label.replace('"', '""')[:80]
    return f'=HYPERLINK("{safe_path}","{safe_label}")'

# Diff PDFs already on disk in docs/rule-pdfs/rule-NN.pdf
def diff_pdf_path(rule_num):
    p = os.path.join(PDFS, f"rule-{rule_num:02d}.pdf")
    return p if os.path.exists(p) else None

NOT_100_GET_TEMPLATE = True

for row_idx, row in enumerate(ROWS, start=2):
    rule_num, section, role, profile, when, what_we_emit, sure_orig, ex_job, ex_plan, arch_key, specific_q, unlocks = row
    long = LONG.get(rule_num, "")
    affected = AFFECTED.get(str(rule_num)) or AFFECTED.get(rule_num)
    affected_str = ""
    if affected:
        affected_str = f"{affected['sticks']:,} sticks / {affected['pairs']} pairs"
    elif sure_orig == "Yes":
        affected_str = "—"

    # 3-state status:
    #   "Yes"    = codec emits correctly today (verified across corpora)
    #   "Spec'd" = rule found in HYTEK config / docs; needs implementation
    #   "No"     = rule still unknown; need Scott's domain answer
    spec = SPEC_FOUND.get(rule_num)
    if spec is not None:
        status = "Spec'd"
        source_text = f"{spec[0]}\n\n{spec[1]}"
    else:
        status = sure_orig
        source_text = "" if sure_orig == "Yes" else ""

    correction_seed = TEMPLATE if (status == "No" and NOT_100_GET_TEMPLATE) else ""

    show_pdf_links = (status != "Yes")
    diff_pdf = diff_pdf_path(rule_num) if show_pdf_links else None
    manuf = manuf_pdf(ex_job, ex_plan) if (ex_job and ex_plan and show_pdf_links) else None
    arch = arch_pdf_path(ex_job, arch_key) if (ex_job and arch_key and show_pdf_links) else None
    fcp = fcp_path(ex_job) if (ex_job and show_pdf_links) else None
    xml = xml_path(ex_job, ex_plan) if (ex_job and ex_plan and show_pdf_links) else None

    # Build the lookup label
    lookup_label = ""
    loc = EXAMPLE_LOC.get(rule_num)
    if show_pdf_links and loc and ex_job and ex_plan:
        frame, stick = loc
        lookup_label = f"{ex_job}\n{ex_plan}\nframe {frame}\nstick {stick}"

    empirical = empirical_text(rule_num)

    cells = [
        rule_num, section, role, profile, when, what_we_emit, status,
        source_text,
        empirical,
        lookup_label,
        specific_q if status == "No" else ("(see Source col)" if status == "Spec'd" else ""),
        affected_str,
        unlocks if status != "Yes" else "—",
        long if status == "No" else "",
        correction_seed,
        hyperlink(diff_pdf, f"rule-{rule_num:02d}.pdf") if diff_pdf else "",
        hyperlink(manuf, os.path.basename(manuf)[:50]) if manuf else "",
        hyperlink(arch, os.path.basename(arch)[:50]) if arch else "",
        hyperlink(fcp, os.path.basename(fcp)[:50]) if fcp else "",
        hyperlink(xml, os.path.basename(xml)[:50]) if xml else "",
    ]
    for col_idx, val in enumerate(cells, start=1):
        c = ws.cell(row=row_idx, column=col_idx, value=val)
        c.font = Font(name="Arial", size=10)
        c.alignment = align_wrap
        c.border = border
    # Color the Status cell
    sure_cell = ws.cell(row=row_idx, column=7)
    sure_cell.alignment = align_center
    if status == "Yes":
        sure_cell.fill = fill_yes
        sure_cell.font = Font(name="Arial", size=10, bold=True, color="2E7D32")
        ws.row_dimensions[row_idx].height = 22
    elif status == "Spec'd":
        sure_cell.fill = fill_specd
        sure_cell.font = Font(name="Arial", size=10, bold=True, color="E65100")
        # Make the lookup cell stand out
        if lookup_label:
            lookup_cell = ws.cell(row=row_idx, column=10)
            lookup_cell.font = Font(name="Arial", size=10, bold=True, color="231F20")
            lookup_cell.fill = PatternFill("solid", start_color="FFF3CD")
        # Highlight empirical column if data present
        if empirical:
            ws.cell(row=row_idx, column=9).fill = PatternFill("solid", start_color="E1F5FE")
        line_count = source_text.count("\n") + max(1, len(source_text) // 70) + empirical.count("\n")
        ws.row_dimensions[row_idx].height = max(80, min(line_count * 14, 260))
    else:  # No
        sure_cell.fill = fill_no
        sure_cell.font = Font(name="Arial", size=10, bold=True, color="C62828")
        # Highlight correction cell (now col 15 after Source + Empirical + Lookup)
        ws.cell(row=row_idx, column=15).fill = fill_correction
        if lookup_label:
            lookup_cell = ws.cell(row=row_idx, column=10)
            lookup_cell.font = Font(name="Arial", size=10, bold=True, color="231F20")
            lookup_cell.fill = PatternFill("solid", start_color="FFF3CD")
        if empirical:
            ws.cell(row=row_idx, column=9).fill = PatternFill("solid", start_color="E1F5FE")
        line_count = max(long.count("\n"), empirical.count("\n")) + max(1, len(long) // 90)
        ws.row_dimensions[row_idx].height = max(140, min(line_count * 13, 380))

    # Make hyperlink cells render as blue underlined (cols 16-20 now)
    for col in (16, 17, 18, 19, 20):
        cell = ws.cell(row=row_idx, column=col)
        if cell.value:
            cell.font = font_link

widths = {1: 5, 2: 9, 3: 26, 4: 11, 5: 32, 6: 42, 7: 10,
          8: 55,  # Source / what to implement
          9: 60,  # Empirical pattern
          10: 22,  # Look up in Detailer
          11: 50, 12: 18, 13: 22, 14: 65, 15: 50,
          16: 22, 17: 30, 18: 30, 19: 30, 20: 30}
for col, w in widths.items():
    ws.column_dimensions[get_column_letter(col)].width = w

ws.freeze_panes = "B2"

# Summary sheet
ws2 = wb.create_sheet("Summary")
ws2["A1"] = "Codec rule review — V2"
ws2["A1"].font = Font(name="Arial", bold=True, size=14)

ws2["A3"] = "Total rules"
ws2["B3"] = f'=COUNTA(\'Codec rules\'!A2:A{1 + len(ROWS)})'

ws2["A4"] = "Yes — codec emits correctly today"
ws2["B4"] = f'=COUNTIF(\'Codec rules\'!G2:G{1 + len(ROWS)},"Yes")'
ws2["B4"].font = Font(name="Arial", bold=True, color="2E7D32")

ws2["A5"] = "Spec'd — rule found, needs implementation"
ws2["B5"] = f"=COUNTIF('Codec rules'!G2:G{1 + len(ROWS)},\"Spec'd\")"
ws2["B5"].font = Font(name="Arial", bold=True, color="E65100")

ws2["A6"] = "No — rule unknown, need Scott"
ws2["B6"] = f'=COUNTIF(\'Codec rules\'!G2:G{1 + len(ROWS)},"No")'
ws2["B6"].font = Font(name="Arial", bold=True, color="C62828")

ws2["A8"] = "Columns explained"
ws2["A8"].font = Font(name="Arial", bold=True, size=12)
explainers = [
    ("Status", "Yes (green) = working today. Spec'd (orange) = rule found, codec needs the implementation. No (red) = rule still unknown, needs Scott."),
    ("Source / what to implement", "For Spec'd rows: cites the source file (machine-setups.ts, fc-dat-rules.ts, design doc) and describes what code change is needed."),
    ("Look up in Detailer", "Job + plan + frame label + stick name. Use this to navigate to the same stick in Detailer or in the Manufacturing PDF."),
    ("Specific question", "The bottom-line question I need answered, lifted out of the long-form text."),
    ("Affected (sticks / pairs)", "How many sticks across the 388-pair corpus this rule's gate would fire on. High count = high stakes."),
    ("Unlocks if fixed", "Approximate parity gain if the rule lands correctly. 'small' / 'moderate' / specific pp where known."),
    ("My uncertainty (long-form)", "Full context: what we emit, what Detailer does, my hypothesis, what I need."),
    ("Scott's correction", "Yellow cell with template scaffold — fill the blanks, that's enough for me to update the codec."),
    ("Diff PDF", "Codec emit (top, red = wrong-emit) vs Detailer reference (bottom, red = miss) for the example stick."),
    ("Manufacturing PDF", "FrameCAD-generated frame-elevation PDF for the example plan. Multi-page; scroll to the cited frame."),
    ("Architectural PDF", "Project plans (LBW MARKUP, SLAB DESIGN, CONSTRUCTION DETAILS, AC LAYOUT) — building context."),
    (".fcp file", "Click to open the Detailer project. Navigate to the cited frame for live tooling."),
    ("Source XML", "Raw XML the codec receives. Inspect to see what input data we have."),
]
for i, (col, desc) in enumerate(explainers):
    ws2.cell(row=9 + i, column=1, value=col).font = Font(name="Arial", bold=True)
    ws2.cell(row=9 + i, column=2, value=desc)
ws2.column_dimensions["A"].width = 35
ws2.column_dimensions["B"].width = 110

ws2.cell(row=23, column=1, value="How to use this workbook").font = Font(name="Arial", bold=True, size=12)
ws2.cell(row=24, column=1, value="1. Filter Status column to 'No' (red) — those are the rules where I genuinely need your domain knowledge.")
ws2.cell(row=25, column=1, value="2. For each No row: click the Diff PDF / Manufacturing PDF / Architectural PDF links to see the issue in context.")
ws2.cell(row=26, column=1, value="3. Read the Specific question — most are answerable in 1-2 lines.")
ws2.cell(row=27, column=1, value="4. Fill the yellow correction cell using the template scaffold.")
ws2.cell(row=28, column=1, value="5. Spec'd rows (orange) are already answered — read Source col if you want the detail. Claude implements them next.")
ws2.cell(row=29, column=1, value="")
ws2.cell(row=30, column=1, value="Path to 90%+ codec parity").font = Font(name="Arial", bold=True, size=12)
ws2.cell(row=31, column=1, value="Current codec parity: 77.46% across 388-pair corpus.")
ws2.cell(row=32, column=1, value="Implementing the 15 Spec'd rules (orange): estimated +15-25pp -> 90%+.")
ws2.cell(row=33, column=1, value="Answering the 25 No rules (red): closes the rest.")

OUT = os.path.join(ROOT, "docs", "rules-review-v4.xlsx")
wb.save(OUT)
print(f"Wrote: {OUT}")
print(f"Rules: {len(ROWS)}")
yes = sum(1 for r in ROWS if r[6] == "Yes")
print(f"100% sure: {yes} / Not 100%: {len(ROWS) - yes}")

# Sanity-check link counts
diff_count = sum(1 for r in ROWS if r[6] == "No" and diff_pdf_path(r[0]))
manuf_count = sum(1 for r in ROWS if r[6] == "No" and manuf_pdf(r[7], r[8]))
arch_count = sum(1 for r in ROWS if r[6] == "No" and arch_pdf_path(r[7], r[9]))
fcp_count = sum(1 for r in ROWS if r[6] == "No" and fcp_path(r[7]))
xml_count = sum(1 for r in ROWS if r[6] == "No" and xml_path(r[7], r[8]))
print(f"Hyperlinks set on No rows:  diff={diff_count}  manuf={manuf_count}  arch={arch_count}  fcp={fcp_count}  xml={xml_count}")
