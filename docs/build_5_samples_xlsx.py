"""Build docs/5-samples-for-scott.xlsx — same 5 reference samples, but
in spreadsheet form with clickable hyperlinks to Detailer + manufacturing PDFs.
"""
import json, os
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
DIFF_DIR = ROOT / "scripts" / "baselines" / "raw-y-pairs"
LINKS_FILE = ROOT / "docs" / "rule-pdfs" / "_links.json"
ARCH_FILE = ROOT / "docs" / "rule-pdfs" / "_arch_pdfs.json"

with LINKS_FILE.open(encoding="utf-8") as f:
    LINKS = json.load(f)
with ARCH_FILE.open(encoding="utf-8") as f:
    ARCH = json.load(f)


def manuf_pdf_for(job, plan):
    return LINKS.get("manufacturing_pdf", {}).get(job, {}).get(plan)


def fcp_for(job):
    return LINKS.get("fcp", {}).get(job)


def xml_for(job, plan):
    return LINKS.get("xml", {}).get(job, {}).get(plan)


def arch_for(job, key="construction_details"):
    if not job: return None
    pdfs = ARCH.get(job, {})
    # Prefer the LBW markup or construction details
    for k in ("lbw_markup", "construction_details", "construction_plans", "carpenter", "ac_layout"):
        if k in pdfs:
            return pdfs[k]
    return None


def get_stick(pair_id, frame_name, stick_name):
    f = DIFF_DIR / f"{pair_id}.json"
    if not f.exists():
        return None
    d = json.loads(f.read_text(encoding="utf-8"))
    for fr in d.get("byFrame", []):
        if fr.get("name") != frame_name:
            continue
        for s in fr.get("sticks", []):
            if s.get("name") == stick_name:
                return s
    return None


def hyperlink(path, label):
    if not path:
        return ""
    uri = "file:///" + str(path).replace("\\", "/")
    safe_path = uri.replace('"', '""')
    safe_label = (label or "open").replace('"', '""')[:80]
    return f'=HYPERLINK("{safe_path}","{safe_label}")'


SAMPLES = [
    {
        "n": 1,
        "title": "NLBW Raised B-plate slab anchors",
        "pair": "HG260001__GF-NLBW-70.075",
        "frame": "N14",
        "stick": "B1",
        "question": "Why does this NLBW raised B-plate (sill above an opening) get slab anchor bolts when an LBW raised B doesn't? What's structurally different?",
        "context": "Raised B-plate (Bh) at z=61.5mm — sill above a door opening, NOT on the slab. Detailer's ref shows Web@8 + Bolt@62. LBW raised B-plates (e.g. PK4 L4 B2) at the SAME elevation do NOT have these anchors. Looking for the rule that fires anchors here.",
    },
    {
        "n": 2,
        "title": "Truss panel-point dimples (sign discriminator)",
        "pair": "HG250011__GF-TIN-70.075",
        "frame": "PC1-1",
        "stick": "T2",
        "question": "Each dimple PAIR (e.g. @562 + @613, @1166 + @1217) sits offset to ONE SIDE of where a web stick projects onto this chord by ~25mm. What determines which side — heel/apex, frame centerline side, web slope direction, or something else?",
        "context": "T2 is a 2.4m TopChord. Ref has multiple InnerDimple PAIRS in body (panel-point markers). Pairs are spaced ~50mm apart. Each pair sits to one side of the web-projection point. Sign flips bimodally per the corpus mining (peaks at ±25mm).",
    },
    {
        "n": 3,
        "title": "Truss chord cap notches (start + end ops)",
        "pair": "HG250011__GF-TIN-70.075",
        "frame": "PC1-1",
        "stick": "T2",
        "question": "When I said 'cap notch' I meant the InnerNotch+LipNotch span at the very start (0..39) and very end (length-39..length). ~80% of TIN chord sticks have these caps; ~20% only have LipNotches without InnerNotches. What physically distinguishes a chord that gets the InnerNotch cap vs one that doesn't?",
        "context": "Same T2 stick as #2. Looking at start/end ops. The InnerNotch+LipNotch is for chord segments that connect to another segment (heel, apex, mid-span splice). The 20% without may be free ends or segments connecting to a different part type.",
    },
    {
        "n": 4,
        "title": "Br/R sticks (decoded with HYTEK codes)",
        "pair": "HG250011__GF-TIN-70.095",
        "frame": "TGI1-1",
        "stick": "R6",
        "question": "What is an R6 stick? Is it a ribbon, lateral brace, ridge, or something else? Confirm whether the 41mm Swage cap + 11mm dimple offset is correct for HYTEK Br/R sticks (vs studs' 39 + 16.5).",
        "context": "R6 is 399mm long. Ref ops: Chamfer@start, ID@99.5, ID@198.5, ID@281.5 (3 dimples in body), LipNotch 77..122, InnerNotch+LipNotch 176..221, InnerNotch 259..304. Codec only emits Swage 357.9..398.9. Want to identify what type of stick this is in HYTEK terminology + confirm cap/dimple constants.",
    },
    {
        "n": 5,
        "title": "Long T plate with InnerNotch in body",
        "pair": "HG250085__GF-TIN-70.095",
        "frame": "TN4-1",
        "stick": "T4",
        "question": "This 4190mm long T plate has an InnerNotch span 192..279 in the BODY (not at ends). What triggers this — above an opening? At a king-stud crossing? At a plate joint? Where in this frame would I look to find the cause?",
        "context": "T4 is a 4190mm top chord. Ref has standard caps at 0..39 and 4151..4190 PLUS an InnerNotch span 192..279 in the body. Most long T plates have ONLY cap notches; ~5% have a body notch like this. Looking for the geometric or structural trigger.",
    },
]

# ----- workbook -----
wb = Workbook()
ws = wb.active
ws.title = "5 samples"

HEADERS = [
    "#",
    "Sample",
    "Job",
    "Plan",
    "Frame label",
    "Stick name",
    "Stick length",
    "My question",
    "Context",
    "Codec emits (extras vs ref)",
    "Detailer reference (missing from codec)",
    "Manufacturing PDF (frame elevation)",
    "Architectural PDF (project context)",
    ".fcp file (open in Detailer)",
    "Source XML",
    "Scott's answer",
]
ws.append(HEADERS)

font_header = Font(name="Arial", bold=True, color="FFFFFF", size=11)
fill_header = PatternFill("solid", start_color="231F20")
align_header = Alignment(horizontal="center", vertical="center", wrap_text=True)
fill_yellow = PatternFill("solid", start_color="FFF9C4")
fill_lookup = PatternFill("solid", start_color="FFF3CD")
font_link = Font(name="Arial", size=10, color="1976D2", underline="single")
align_wrap = Alignment(vertical="top", wrap_text=True)
align_center = Alignment(horizontal="center", vertical="top")
thin = Side(border_style="thin", color="CCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for cell in ws[1]:
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = align_header
    cell.border = border
ws.row_dimensions[1].height = 36

for i, s in enumerate(SAMPLES, start=2):
    job = s["pair"].split("__")[0]
    plan = s["pair"].split("__")[1]
    stick = get_stick(s["pair"], s["frame"], s["stick"])

    if stick:
        L = stick.get("refLength") or stick.get("oursLength") or 0
        L_str = f"{L:.1f} mm"
        extras = stick.get("extras", []) or []
        missing = stick.get("missing", []) or []
        extras_str = "\n".join(extras[:25]) if extras else "(no extras — codec aligned)"
        missing_str = "\n".join(missing[:25]) if missing else "(no missing — codec matches ref)"
    else:
        L_str = "?"
        extras_str = "(stick not found in diff data)"
        missing_str = "(stick not found in diff data)"

    manuf = manuf_pdf_for(job, plan)
    arch = arch_for(job)
    fcp = fcp_for(job)
    xml = xml_for(job, plan)

    cells = [
        s["n"],
        s["title"],
        job,
        plan,
        s["frame"],
        s["stick"],
        L_str,
        s["question"],
        s["context"],
        extras_str,
        missing_str,
        hyperlink(manuf, os.path.basename(manuf)[:60]) if manuf else "",
        hyperlink(arch, os.path.basename(arch)[:60]) if arch else "",
        hyperlink(fcp, os.path.basename(fcp)[:60]) if fcp else "",
        hyperlink(xml, os.path.basename(xml)[:60]) if xml else "",
        "",  # Scott's answer column - blank
    ]

    for col_idx, val in enumerate(cells, start=1):
        c = ws.cell(row=i, column=col_idx, value=val)
        c.font = Font(name="Arial", size=10)
        c.alignment = align_wrap
        c.border = border

    # Highlight key cells
    ws.cell(row=i, column=5).fill = fill_lookup  # Frame label
    ws.cell(row=i, column=6).fill = fill_lookup  # Stick name
    ws.cell(row=i, column=5).font = Font(name="Arial", size=10, bold=True)
    ws.cell(row=i, column=6).font = Font(name="Arial", size=10, bold=True)

    # Yellow Scott's answer cell
    ws.cell(row=i, column=16).fill = fill_yellow

    # Hyperlinks render blue + underline
    for col in (12, 13, 14, 15):
        cell = ws.cell(row=i, column=col)
        if cell.value:
            cell.font = font_link

    # Tall row to fit ops + question + context
    ws.row_dimensions[i].height = 280

# Column widths
widths = {1: 4, 2: 36, 3: 11, 4: 22, 5: 12, 6: 10, 7: 12,
          8: 50, 9: 50,
          10: 36, 11: 36,
          12: 32, 13: 28, 14: 28, 15: 28,
          16: 60}
for col, w in widths.items():
    ws.column_dimensions[get_column_letter(col)].width = w

ws.freeze_panes = "B2"

# Summary sheet
ws2 = wb.create_sheet("How to use")
ws2["A1"] = "5 reference samples — how to use"
ws2["A1"].font = Font(name="Arial", bold=True, size=14)
ws2["A3"] = "1. Click the Manufacturing PDF link to open the per-plan frame-elevation drawings."
ws2["A4"] = "2. Use the Frame label (column E) to find the right page in the PDF."
ws2["A5"] = "3. Use the Stick name (column F) to identify the specific stick on that elevation."
ws2["A6"] = "4. Compare 'Codec emits' (column J) vs 'Detailer reference' (column K) for the gap."
ws2["A7"] = "5. Click the .fcp link (column N) to open the same project in Detailer for live inspection."
ws2["A8"] = "6. Type your answer in the yellow 'Scott's answer' column (P)."
ws2["A9"] = "7. Save and hand back to Claude — the codec will be updated."

ws2.column_dimensions["A"].width = 100

OUT = ROOT / "docs" / "5-samples-for-scott.xlsx"
wb.save(OUT)
print(f"Wrote: {OUT}")
print(f"Samples: {len(SAMPLES)}")
